
import json
import sys
import subprocess

# discord token
TOKEN = ''

# API token(s)
ALCHEMY = ""
OPENSEA = ""

# keep this up to date if a new collection gets released
collections = {
    "furniture": "0xec4de0a00c694cc7957fb90b9005b24a3f4f8b99",
    "clothes": "0x8c15d753c4336617890ff9e82c88aa047762b867",
    "pets": "0x792df6705032cd3ad8a6aa3b3b7b0a42c0b9759c",
    "addons": "0xacc8b12fd8b08ecea19fb586c0c744f423fc3dd2",
    "tokens": "0xfbf1c1c09a94fe45ea8cc981c478816963ec958c"
}


# dependencies
def install_dependecies():
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'discord'])
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'datetime'])
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'requests'])
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    subprocess.check_call(
        [sys.executable, '-m', 'pip', 'install', 'excel2img'])
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'random'])
    import discord
    from discord.ext import commands, tasks
    from datetime import datetime
    import requests
    import time
    import openpyxl
    from openpyxl.styles import PatternFill
    import excel2img
    import random


try:
    import discord
    from discord.ext import commands, tasks
    from datetime import datetime
    import requests
    import time
    import openpyxl
    from openpyxl.styles import PatternFill
    import excel2img
    import random
except:
    install_dependecies()


intents = discord.Intents().all()
client = commands.Bot(command_prefix=',', intents=intents)

# loop while bot is running
@tasks.loop(seconds=600)
async def ppcloop():
    total = 0
    alllisted = 0
    emeraldamounts = ['Emerald Stone (10 Emeralds)', 'Emerald Bowl (50 Emeralds)', '100 Carat Emerald', 'Sack of 500 Emeralds',
                      '1000 Carat Emerald', 'Emerald Dino Egg (2,000 Emeralds)', 'Chest of 10k Emeralds']
    ethprice = requests.get(
        'https://min-api.cryptocompare.com/data/price?fsym=ETH&tsyms=USD').json()
    eth = ethprice['USD']
    for emeraldamount in emeraldamounts:
        r = requests.get('https://api.x.immutable.com/v3/orders?auxiliary_fee_percentages=1&auxiliary_fee_recipients=0x12cB8E42c7ec27d30df6Cb8f44aa6445D0e1a78C&buy_token_type=ETH&direction=asc&include_fees=true&order_by=buy_quantity_with_fees&page_size=48&sell_token_address=0xfbf1c1c09a94fe45ea8cc981c478816963ec958c&sell_token_type=ERC721&status=active&sell_token_name='+emeraldamount).json()
        for results in r['result']:
            if emeraldamount == results['sell']['data']['properties']['name']:
                pricequantity = results['buy']['data']['quantity_with_fees']
                pricedecimal = results['buy']['data']['decimals']
                price = int(pricequantity) / (10 ** int(pricedecimal))
                break
        USDprice = price * eth
        if emeraldamount == 'Chest of 10k Emeralds':
            tokens = 10000
        else:
            tokens = int(''.join(filter(str.isdigit, emeraldamount)))
        ppc = round(USDprice/tokens, 3)
        total = total + ppc
        alllisted = alllisted + 1
    total = (total/alllisted)
    await client.change_presence(activity=discord.Game(name="PPC: " + (str(round(total, 3)))))


@client.event
async def on_message(message):
    # HELP
    if message.content.lower() == ',help':
        embed = discord.Embed(title=",help", description="**commands:**")
        embed.add_field(name=",cheapest (furniture/clothes/pets/addons)",
                        value="Shows the cheapest item of a collection on immutable.", inline=False)
        embed.add_field(name=",balance (address)",
                        value="Shows the unminted NFT credits of any address.", inline=False)
        embed.add_field(name=",level (address)",
                        value="Shows the level + XP of any address.", inline=False)
        embed.add_field(name=",search (collection name) (item name)",
                        value="Gives you the ETH price of a specific furni or clothing item.", inline=False)
        embed.add_field(
            name=",ppc", value="Shows you the Price Per Credit for Emeralds.", inline=False)
        embed.add_field(
            name=",ppn", value="Shows you the Price Per Credit for NFTC.", inline=False)
        embed.add_field(name=",calculate (daily nft credits)",
                        value="Calculates your daily + yearly dollar income.", inline=False)
        embed.add_field(
            name=",apr", value="Yearly percentage return on investment. (excluding converting and imx fees)", inline=False)
        embed.add_field(
            name=",giftbox (giftbox code name)", value="Will give you a fair value calculation of a giftbox.", inline=False
        )
        embed.add_field(
            name=",nftc", value="Gets the NFTc picture.", inline=False
        )
        embed.add_field(
            name=",factory", value="Gets the factory pictures.", inline=False
        )
        embed.add_field(
            name=",event", value="Gets the event pictures.", inline=False
        )
        embed.add_field(
            name=",crate", value="Gets the crate options.", inline=False
        )
        await message.channel.send(embed=embed)

    # UHELP
    if message.content.lower() == ',uhelp':
        embed = discord.Embed(title=",uhelp", description="**Unlisted Help:**")
        embed.add_field(name=",pl",
                        value="Creates the P/L files and puts them in chat.", inline=False)
        embed.add_field(name=",nickreset",
                        value="Just resets the nickname to Dark Habbo. Was experimenting with this and left it in.", inline=False)
        embed.add_field(name=",json",
                        value="Creates and downloads a json of all items.", inline=False)
        embed.add_field(name="**mystery command**",
                        value=",dylan", inline=False)
        await message.channel.send(embed=embed)

    # cheapest
    if message.content.lower().startswith(',cheapest'):
        print(message.content)
        try:
            collection = message.content.split(' ', 1)[1]
        except:
            collection = 'shitcode'

        if collection in collections:
            collectionadd = collections[collection]

            r = requests.get('https://api.x.immutable.com/v3/orders?auxiliary_fee_percentages=1&auxiliary_fee_recipients=0x12cB8E42c7ec27d30df6Cb8f44aa6445D0e1a78C&buy_token_type=ETH&direction=asc&include_fees=true&order_by=buy_quantity_with_fees&page_size=48&sell_token_address=' +
                             collectionadd+'&sell_token_type=ERC721&status=active').json()
            orderid = r['result'][0]['order_id']
            productname = r['result'][0]['sell']['data']['properties']['name']
            image = r['result'][0]['sell']['data']['properties']['image_url']
            tokenid = r['result'][0]['sell']['data']['token_id']
            urlcode = r['result'][0]['sell']['data']['token_id']
            pricequantity = r['result'][0]['buy']['data']['quantity_with_fees']
            pricedecimal = r['result'][0]['buy']['data']['decimals']
            price = int(pricequantity) / (10 ** int(pricedecimal))
            r = requests.get(
                'https://api.x.immutable.com/v1/assets/'+collectionadd+'/'+tokenid).json()
            score = (str(r['metadata']['score']))
            r = requests.get(
                'https://min-api.cryptocompare.com/data/price?fsym=ETH&tsyms=USD').json()
            eth = r['USD']
            USDprice = price * eth
            # embed = discord.Embed(title=productname, url='https://market.immutable.com/collections/'+collectionadd+'/assets/'+urlcode)
            productlinkname = productname.replace(' ', '%20')
            embed = discord.Embed(
                title=productname, url='https://furni.app/furni/'+productlinkname)
            embed.add_field(name="ETH:", value=price, inline=True)
            embed.add_field(name="USD:", value=round(USDprice, 2), inline=True)
            embed.add_field(name="XP SCORE:", value=score, inline=False)
            embed.add_field(
                name=' ', value='[Buy](https://furni.app/checkout/'+(str(orderid))+')', inline=False)
            embed.set_thumbnail(url=image)
            await message.channel.send(embed=embed)
        else:
            collectionstring = ''
            for collectionitems in collections:
                collectionstring += collectionitems + ' \n'
            await message.channel.send('**Use one of these:** \n'+collectionstring)

    # balance
    if message.content.lower().startswith(',balance'):
        print(message.content)
        address = message.content.split(' ', 1)[1]
        if address == "dylan":
            address = '0xadc493644308c46a1f91f1b5aec8d2b7f14aa8aa'
        if ".eth" in address:
            r = requests.get(
                'https://api.ensideas.com/ens/resolve/'+address).json()
            try:
                address = r['address']
            except:
                await message.channel.send("You tried an invalid ENS")
        r = requests.get(
            'https://collectibles.habbo.com/api/tokens/balance/?address='+address).json()
        balance = r['balance']
        buffer = r['buffer']['credits']
        days = r['buffer']['counter']
        daysleft = 60-days

        r = requests.get(
            'https://api.x.immutable.com/v2/balances/'+address).json()
        for results in r['result']:
            if results['symbol'] == "ETH":
                imxeth = round(int(results['balance']) / (10 ** 18), 4)
                r = requests.get(
                    'https://min-api.cryptocompare.com/data/price?fsym=ETH&tsyms=USD').json()
                eth = r['USD']
                imxethdollar = round(imxeth * eth, 2)

        minbal = 0
        minem = 0
        r = requests.get(
            'https://api.x.immutable.com/v1/assets?collection=0xFbf1C1C09a94FE45EA8CC981c478816963eC958c&user='+address).json()
        for credits in r['result']:
            name = credits['metadata']['name']
            if 'emerald' in name.lower():
                if name == 'Chest of 10k Emeralds':
                    tokens = 10000
                elif name == 'Emerald Dino Egg (2,000 Emeralds)':
                    tokens = 2000
                else:
                    digit_str = ''.join(filter(str.isdigit, name.replace(
                        '(', '').replace(')', '').replace(',', '')))
                    tokens = int(digit_str) if digit_str else 0
                minem += int(tokens)
            elif 'mint' in name.lower():
                # fuck you mint token
                pass
            else:
                value = credits['metadata']['type'].replace(
                    'CREDIT', '').replace('NFTC', '').replace('nft_emerald_', '')
                minbal += int(value)

        await message.channel.send("Current unminted balance: " + (str(balance)) + ' emeralds\nCurrent minted balance: '
                                   + (str(minbal)) + ' nftc\nCurrent minted balance: ' + (str(minem)) + ' emeralds\nCurrent unclaimed balance: ' + (str(buffer)) + ' emeralds ('+(str(daysleft))+' days left)\n\nL2 ETH: ' + (str(imxeth)) + ' ($' + (str(imxethdollar))+')')

    # search
    if message.content.lower().startswith(',search'):
        print(message.content)
        try:
            collection = message.content.split(' ', 2)[1]
            query = message.content.split(' ', 2)[2].lower()
            query_tokens = set(query.split())
        except:
            collection = 'shitcode'
            query = 'shitcode'

        if collection in collections:
            collectionadd = collections[collection]

            try:
                r = requests.get('https://api.x.immutable.com/v3/orders?auxiliary_fee_percentages=1&auxiliary_fee_recipients=0x12cB8E42c7ec27d30df6Cb8f44aa6445D0e1a78C&buy_token_type=ETH&direction=asc&include_fees=true&order_by=buy_quantity_with_fees&page_size=200&sell_token_address=' +
                                 collectionadd+'&sell_token_type=ERC721&status=active&sell_token_name='+query).json()
                hit = False
                for results in r['result']:
                    if query == results['sell']['data']['properties']['name'].lower():
                        orderid = results['order_id']
                        productname = results['sell']['data']['properties']['name']
                        image = results['sell']['data']['properties']['image_url']
                        tokenid = r['result'][0]['sell']['data']['token_id']
                        urlcode = results['sell']['data']['token_id']
                        pricequantity = results['buy']['data']['quantity_with_fees']
                        pricedecimal = results['buy']['data']['decimals']
                        price = int(pricequantity) / (10 ** int(pricedecimal))
                        r = requests.get(
                            'https://api.x.immutable.com/v1/assets/'+collectionadd+'/'+tokenid).json()
                        try:
                            score = (str(r['metadata']['score']))
                        except:
                            score = 'error'
                        r = requests.get(
                            'https://min-api.cryptocompare.com/data/price?fsym=ETH&tsyms=USD').json()
                        eth = r['USD']
                        USDprice = price * eth
                        hit = True
                        # embed = discord.Embed(title=productname, url='https://market.immutable.com/collections/'+collectionadd+'/assets/'+urlcode)
                        productlinkname = productname.replace(' ', '%20')
                        embed = discord.Embed(
                            title=productname, url='https://furni.app/furni/'+productlinkname)
                        embed.add_field(name="ETH:", value=price, inline=True)
                        embed.add_field(name="USD:", value=round(
                            USDprice, 2), inline=True)
                        embed.add_field(name="XP SCORE:",
                                        value=score, inline=False)
                        embed.set_thumbnail(url=image)
                        embed.add_field(
                            name=' ', value='[Buy](https://furni.app/checkout/'+(str(orderid))+')', inline=False)
                        await message.channel.send(embed=embed)
                        break

                if hit == False:
                    for results in r['result']:
                        if query_tokens.issubset(set(results['sell']['data']['properties']['name'].lower().split())) or query in query in results['sell']['data']['properties']['name'].lower():
                            orderid = results['order_id']
                            productname = results['sell']['data']['properties']['name']
                            image = results['sell']['data']['properties']['image_url']
                            tokenid = r['result'][0]['sell']['data']['token_id']
                            urlcode = results['sell']['data']['token_id']
                            pricequantity = results['buy']['data']['quantity_with_fees']
                            pricedecimal = results['buy']['data']['decimals']
                            price = int(pricequantity) / \
                                (10 ** int(pricedecimal))
                            r = requests.get(
                                'https://api.x.immutable.com/v1/assets/'+collectionadd+'/'+tokenid).json()
                            try:
                                score = (str(r['metadata']['score']))
                            except:
                                score = 'error'
                            r = requests.get(
                                'https://min-api.cryptocompare.com/data/price?fsym=ETH&tsyms=USD').json()
                            eth = r['USD']
                            USDprice = price * eth
                            hit = True
                            # embed = discord.Embed(title=productname, url='https://market.immutable.com/collections/'+collectionadd+'/assets/'+urlcode)
                            productlinkname = productname.replace(' ', '%20')
                            embed = discord.Embed(
                                title=productname, url='https://furni.app/furni/'+productlinkname)
                            embed.add_field(
                                name="ETH:", value=price, inline=True)
                            embed.add_field(name="USD:", value=round(
                                USDprice, 2), inline=True)
                            embed.add_field(name="XP SCORE:",
                                            value=score, inline=False)
                            embed.set_thumbnail(url=image)
                            embed.add_field(
                                name=' ', value='[Buy](https://furni.app/checkout/'+(str(orderid))+')', inline=False)
                            await message.channel.send(embed=embed)
                            break

            except Exception as e:
                print(e)

        else:
            collectionstring = ''
            for collectionitems in collections:
                collectionstring += collectionitems + ' \n'
            await message.channel.send('**Use one of these:** \n'+collectionstring)

    # PPC
    if message.content.lower() == ',ppc' or message.content.lower() == ',ppe':
        print(message.content)
        total = 0
        alllisted = 0
        # tokenamounts = ['10 NFTC', '50 NFTC', '100 NFTC', '500 NFTC',
        #                 '1000 NFTC', '10000 NFTC', '50000 NFTC', '100000 NFTC']
        emeraldamounts = ['Emerald Stone (10 Emeralds)', 'Emerald Bowl (50 Emeralds)', '100 Carat Emerald', 'Sack of 500 Emeralds',
                          '1000 Carat Emerald', 'Emerald Dino Egg (2,000 Emeralds)', 'Chest of 10k Emeralds']
        ethprice = requests.get(
            'https://min-api.cryptocompare.com/data/price?fsym=ETH&tsyms=USD').json()
        # NFTC
        # embed = discord.Embed(title='Price Per Credit')
        # for tokenamount in tokenamounts:
        #     r = requests.get('https://api.x.immutable.com/v3/orders?auxiliary_fee_percentages=1&auxiliary_fee_recipients=0x12cB8E42c7ec27d30df6Cb8f44aa6445D0e1a78C&buy_token_type=ETH&direction=asc&include_fees=true&order_by=buy_quantity_with_fees&page_size=48&sell_token_address=0xfbf1c1c09a94fe45ea8cc981c478816963ec958c&sell_token_type=ERC721&status=active&sell_token_name='+tokenamount).json()
        #     for results in r['result']:
        #         if tokenamount == results['sell']['data']['properties']['name']:
        #             productname = results['sell']['data']['properties']['name']
        #             pricequantity = results['buy']['data']['quantity_with_fees']
        #             pricedecimal = results['buy']['data']['decimals']
        #             price = int(pricequantity) / (10 ** int(pricedecimal))

        #             eth = ethprice['USD']
        #             USDprice = price * eth
        #             tokens = int(tokenamount.replace(' NFTC', ''))
        #             ppc = round(USDprice/tokens, 3)
        #             if tokens <= 1000:
        #                 total = total + ppc
        #                 alllisted = alllisted + 1
        #             # embed.add_field(name=productname, value='$'+(str(ppc))+'\n[$'+(str(round(USDprice,2)))+'](https://market.immutable.com/collections/0xfbf1c1c09a94fe45ea8cc981c478816963ec958c?keywordSearch='+(str(tokens))+')', inline=False)
        #             embed.add_field(name=productname, value='$'+(str(ppc))+'\n[$'+(str(round(
        #                 USDprice, 2)))+'](https://furni.app/furni/'+(tokenamount.replace(' ', '-'))+')', inline=False)
        #             break

        # # Only counts amounts 1k and lower
        # total = (total/alllisted)

        # embed.set_footer(text="Average ppc is $" + (str(round(total, 3))))
        # await message.channel.send(embed=embed)

        # Emeralds
        total = 0
        alllisted = 0
        embed = discord.Embed(
            title='Price Per Credit (NFTC has been moved to ,ppn)')
        for emeraldamount in emeraldamounts:
            r = requests.get('https://api.x.immutable.com/v3/orders?auxiliary_fee_percentages=1&auxiliary_fee_recipients=0x12cB8E42c7ec27d30df6Cb8f44aa6445D0e1a78C&buy_token_type=ETH&direction=asc&include_fees=true&order_by=buy_quantity_with_fees&page_size=48&sell_token_address=0xfbf1c1c09a94fe45ea8cc981c478816963ec958c&sell_token_type=ERC721&status=active&sell_token_name='+emeraldamount).json()
            for results in r['result']:
                if emeraldamount == results['sell']['data']['properties']['name']:
                    productname = results['sell']['data']['properties']['name']
                    pricequantity = results['buy']['data']['quantity_with_fees']
                    pricedecimal = results['buy']['data']['decimals']
                    price = int(pricequantity) / (10 ** int(pricedecimal))

                    eth = ethprice['USD']
                    USDprice = price * eth
                    if emeraldamount == 'Chest of 10k Emeralds':
                        tokens = 10000
                    else:
                        tokens = int(
                            ''.join(filter(str.isdigit, emeraldamount)))
                    ppc = round(USDprice/tokens, 3)
                    total = total + ppc
                    alllisted = alllisted + 1
                    # embed.add_field(name=productname, value='$'+(str(ppc))+'\n[$'+(str(round(USDprice,2)))+'](https://market.immutable.com/collections/0xfbf1c1c09a94fe45ea8cc981c478816963ec958c?keywordSearch='+(str(tokens))+')', inline=False)
                    embed.add_field(name=productname, value='$'+(str(ppc))+'\n[$'+(str(round(
                        USDprice, 2)))+'](https://furni.app/furni/'+(emeraldamount.replace(' ', '-'))+')', inline=False)
                    break

        total = (total/alllisted)

        embed.set_footer(text="Average ppc is $" + (str(round(total, 3))))
        await message.channel.send(embed=embed)

    # ,ppn
    if message.content.lower() == ',ppn':
        tokenamounts = ['10 NFTC', '50 NFTC', '100 NFTC', '500 NFTC',
                        '1000 NFTC', '10000 NFTC', '50000 NFTC', '100000 NFTC']
        total = 0
        alllisted = 0
        # tokenamounts = ['10 NFTC', '50 NFTC', '100 NFTC', '500 NFTC',
        #                 '1000 NFTC', '10000 NFTC', '50000 NFTC', '100000 NFTC']
        emeraldamounts = ['Emerald Stone (10 Emeralds)', 'Emerald Bowl (50 Emeralds)', '100 Carat Emerald', 'Sack of 500 Emeralds',
                          '1000 Carat Emerald', 'Emerald Dino Egg (2,000 Emeralds)', 'Chest of 10k Emeralds']
        ethprice = requests.get(
            'https://min-api.cryptocompare.com/data/price?fsym=ETH&tsyms=USD').json()
        embed = discord.Embed(title='Price Per Credit')
        for tokenamount in tokenamounts:
            r = requests.get('https://api.x.immutable.com/v3/orders?auxiliary_fee_percentages=1&auxiliary_fee_recipients=0x12cB8E42c7ec27d30df6Cb8f44aa6445D0e1a78C&buy_token_type=ETH&direction=asc&include_fees=true&order_by=buy_quantity_with_fees&page_size=48&sell_token_address=0xfbf1c1c09a94fe45ea8cc981c478816963ec958c&sell_token_type=ERC721&status=active&sell_token_name='+tokenamount).json()
            for results in r['result']:
                if tokenamount == results['sell']['data']['properties']['name']:
                    productname = results['sell']['data']['properties']['name']
                    pricequantity = results['buy']['data']['quantity_with_fees']
                    pricedecimal = results['buy']['data']['decimals']
                    price = int(pricequantity) / (10 ** int(pricedecimal))

                    eth = ethprice['USD']
                    USDprice = price * eth
                    tokens = int(tokenamount.replace(' NFTC', ''))
                    ppc = round(USDprice/tokens, 3)
                    if tokens <= 1000:
                        total = total + ppc
                        alllisted = alllisted + 1
                    # embed.add_field(name=productname, value='$'+(str(ppc))+'\n[$'+(str(round(USDprice,2)))+'](https://market.immutable.com/collections/0xfbf1c1c09a94fe45ea8cc981c478816963ec958c?keywordSearch='+(str(tokens))+')', inline=False)
                    embed.add_field(name=productname, value='$'+(str(ppc))+'\n[$'+(str(round(
                        USDprice, 2)))+'](https://furni.app/furni/'+(tokenamount.replace(' ', '-'))+')', inline=False)
                    break

        # Only counts amounts 1k and lower
        total = (total/alllisted)

        embed.set_footer(text="Average ppc is $" + (str(round(total, 3))))
        await message.channel.send(embed=embed)

    # calculate
    if message.content.lower().startswith(',calculate'):
        print(message.content)
        dc = message.content.split(' ', 1)[1]
        dc = int(dc)
        emeraldamounts = ['Emerald Stone (10 Emeralds)', 'Emerald Bowl (50 Emeralds)', '100 Carat Emerald', 'Sack of 500 Emeralds',
                          '1000 Carat Emerald', 'Emerald Dino Egg (2,000 Emeralds)', 'Chest of 10k Emeralds']
        ethprice = requests.get(
            'https://min-api.cryptocompare.com/data/price?fsym=ETH&tsyms=USD').json()
        total = 0
        alllisted = 0
        embed = discord.Embed(title='Daily Credit Calculations')
        for emeraldamount in emeraldamounts:
            r = requests.get('https://api.x.immutable.com/v3/orders?auxiliary_fee_percentages=1&auxiliary_fee_recipients=0x12cB8E42c7ec27d30df6Cb8f44aa6445D0e1a78C&buy_token_type=ETH&direction=asc&include_fees=true&order_by=buy_quantity_with_fees&page_size=48&sell_token_address=0xfbf1c1c09a94fe45ea8cc981c478816963ec958c&sell_token_type=ERC721&status=active&sell_token_name='+emeraldamount).json()
            for results in r['result']:
                if emeraldamount == results['sell']['data']['properties']['name']:
                    pricequantity = results['buy']['data']['quantity_with_fees']
                    pricedecimal = results['buy']['data']['decimals']
                    price = int(pricequantity) / (10 ** int(pricedecimal))
                    eth = ethprice['USD']
                    USDprice = price * eth
                    if emeraldamount == 'Chest of 10k Emeralds':
                        tokens = 10000
                    else:
                        tokens = int(
                            ''.join(filter(str.isdigit, emeraldamount)))
                    ppc = round(USDprice/tokens, 3)
                    total = total + ppc
                    alllisted = alllisted + 1
                    break
        total = (total/alllisted)
        embed.add_field(name="Yearly Emeralds", value=(
            str(round(dc * 365.25))), inline=False)
        embed.add_field(name="Daily Dollar Amount", value='$' +
                        (str((round(dc * total, 2)))), inline=False)
        embed.add_field(name="Yearly Dollar Amount", value='$' +
                        (str((round(dc * total * 365.25)))), inline=False)
        embed.set_footer(text="PPC used: $" + (str(round(total, 3))))
        await message.channel.send(embed=embed)

    # apr
    if message.content.lower() == ',apr':
        print(message.content)
        headers = {
            "accept": "application/json",
            "x-api-key": OPENSEA
        }
        avatarfloor = requests.get(
            'https://api.opensea.io/api/v2/listings/collection/habbo-avatars/best', headers=headers).json()
        portraitfloor = requests.get(
            'https://api.opensea.io/api/v2/listings/collection/habbo-portraits/best', headers=headers).json()
        roomfloor = requests.get(
            'https://api.opensea.io/api/v2/listings/collection/habbo-x-rooms/best', headers=headers).json()
        craftedfloor = requests.get(
            'https://api.opensea.io/api/v2/listings/collection/crafted-avatars/best', headers=headers).json()

        ethprice = requests.get(
            'https://min-api.cryptocompare.com/data/price?fsym=ETH&tsyms=USD').json()

        emeraldamounts = ['Emerald Stone (10 Emeralds)', 'Emerald Bowl (50 Emeralds)', '100 Carat Emerald', 'Sack of 500 Emeralds',
                          '1000 Carat Emerald', 'Emerald Dino Egg (2,000 Emeralds)', 'Chest of 10k Emeralds']

        total = 0
        alllisted = 0
        embed = discord.Embed(title='Annual Percentage Rate')
        for emeraldamount in emeraldamounts:
            r = requests.get('https://api.x.immutable.com/v3/orders?auxiliary_fee_percentages=1&auxiliary_fee_recipients=0x12cB8E42c7ec27d30df6Cb8f44aa6445D0e1a78C&buy_token_type=ETH&direction=asc&include_fees=true&order_by=buy_quantity_with_fees&page_size=48&sell_token_address=0xfbf1c1c09a94fe45ea8cc981c478816963ec958c&sell_token_type=ERC721&status=active&sell_token_name='+emeraldamount).json()
            for results in r['result']:
                if emeraldamount == results['sell']['data']['properties']['name']:
                    pricequantity = results['buy']['data']['quantity_with_fees']
                    pricedecimal = results['buy']['data']['decimals']
                    price = int(pricequantity) / (10 ** int(pricedecimal))
                    eth = ethprice['USD']
                    USDprice = price * eth
                    if emeraldamount == 'Chest of 10k Emeralds':
                        tokens = 10000
                    else:
                        tokens = int(
                            ''.join(filter(str.isdigit, emeraldamount)))
                    ppc = round(USDprice/tokens, 3)
                    total = total + ppc
                    alllisted = alllisted + 1
                    break
        total = (total/alllisted)
        if 'success' not in avatarfloor:
            avatarprice = int(
                avatarfloor['listings'][0]['price']['current']['value']) / (10 ** 18) * eth
            avataryield = (12 * total * 365.25) / avatarprice * 100
            embed.add_field(name="<:avatar:1190796398229213306> Avatar Yield", value=('10 emeralds + 2 room emeralds\n' +
                                                                                      str(round(avataryield, 1))) + '% - $' + (str(round(avatarprice, 2))), inline=False)
        if 'success' not in portraitfloor:
            portraitprice = int(
                portraitfloor['listings'][0]['price']['current']['value']) / (10 ** 18) * eth
            portraityield = (5 * total * 365.25) / portraitprice * 100
            embed.add_field(name="<:portrait:1190796399462326282> Portrait Yield", value=('5 emeralds\n' +
                                                                                          str(round(portraityield, 1)))+'% - $' + (str(round(portraitprice, 2))), inline=False)
        if 'success' not in roomfloor:
            roomprice = int(roomfloor['listings'][0]['price']
                            ['current']['value']) / (10 ** 18) * eth
            roomyield = (1 * total * 365.25) / roomprice * 100
            embed.add_field(name="<:rooms:1190796400540274809> Room Yield", value=('1 emerald\n' +
                                                                                   str(round(roomyield, 1)))+'% - $' + (str(round(roomprice, 2))), inline=False)
        if 'success' not in craftedfloor:
            craftedprice = int(
                craftedfloor['listings'][0]['price']['current']['value']) / (10 ** 18) * eth
            craftedyield = (34 * total * 365.25) / craftedprice * 100
            embed.add_field(name="<:craftedavatar:1190796394093629520> Crafted Yield", value=('30 emeralds + 4 room emeralds\n' +
                                                                                              str(round(craftedyield, 1)))+'% - $' + (str(round(craftedprice, 2))) + '\n\n' + '[Avatars](https://opensea.io/collection/habbo-avatars)' + ' | ' + '[Portraits](https://opensea.io/collection/habbo-portraits)' +
                            ' | ' + '[Rooms](https://opensea.io/collection/habbo-x-rooms)' + ' | ' + '[Crafted Avatars](https://opensea.io/collection/crafted-avatars)', inline=False)
        embed.set_footer(text="PPC used: $" + (str(round(total, 3))))
        await message.channel.send(embed=embed)
    # why do I calculate the ppc in every command instead of making a function of it? Because I am too stupid to get that to work.

    # giftbox
    if message.content.lower().startswith(',giftbox'):
        print(message.content)
        giftboxarr = []
        try:
            giftbox = message.content.split(' ', 1)[1].lower()
        except:
            giftbox = 'shitcode'
        fapp = requests.get('https://request.ninja/fapplist').json()
        if '_' not in giftbox:
            query_tokens = set(giftbox.split())
            for giftboxsearch in fapp:
                if giftbox in giftboxsearch['metaDataName'].lower() or query_tokens.issubset(set(giftboxsearch['metaDataName'].lower().split())):
                    giftbox = giftboxsearch['productCode']
                    break

        r = requests.get(
            'https://collectibles.habbo.com/api/giftbox/droprates/').json()
        for giftboxes in r['droprates']:
            giftboxarr.append(giftboxes)
        if giftbox in giftboxarr:
            # fair value x100!!!!
            fair_value = 0
            droprates = r['droprates'][giftbox]
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.cell(row=1, column=1).value = "Name"
            sheet.cell(row=1, column=2).value = "Dollar Price"
            sheet.cell(row=1, column=3).value = "Chance"
            sheet.cell(row=1, column=4).value = "Dollar * Chance"
            sheet['A1'].fill = PatternFill("solid", start_color="c4c2c2")
            sheet['B1'].fill = PatternFill("solid", start_color="c4c2c2")
            sheet['C1'].fill = PatternFill("solid", start_color="c4c2c2")
            sheet['D1'].fill = PatternFill("solid", start_color="c4c2c2")
            counter = 2
            ethprice = requests.get(
                'https://min-api.cryptocompare.com/data/price?fsym=ETH&tsyms=USD').json()
            eth = ethprice['USD']

            for fappsearch in fapp:
                if fappsearch['productCode'] == giftbox:
                    boxname = fappsearch['metaDataName']
            r = requests.get('https://api.x.immutable.com/v3/orders?auxiliary_fee_percentages=1&auxiliary_fee_recipients=0x12cB8E42c7ec27d30df6Cb8f44aa6445D0e1a78C&buy_token_type=ETH&direction=asc&include_fees=true&order_by=buy_quantity_with_fees&page_size=50&sell_token_address=0xec4de0a00c694cc7957fb90b9005b24a3f4f8b99&sell_token_type=ERC721&status=active&sell_token_name='+boxname).json()
            for results in r['result']:
                boxprice = 0
                if boxname == results['sell']['data']['properties']['name']:
                    pricequantity = results['buy']['data']['quantity_with_fees']
                    pricedecimal = results['buy']['data']['decimals']
                    image = results['sell']['data']['properties']['image_url']
                    price = int(pricequantity) / (10 ** int(pricedecimal))
                    boxprice = price * eth
                    break
            cop = 0

            for items in droprates:
                percentage = droprates[items]
                for fappsearch in fapp:
                    if fappsearch['productCode'] == items:
                        itemname = fappsearch['metaDataName']
                        itemcollection = fappsearch['collection']

                        r = requests.get('https://api.x.immutable.com/v3/orders?auxiliary_fee_percentages=1&auxiliary_fee_recipients=0x12cB8E42c7ec27d30df6Cb8f44aa6445D0e1a78C&buy_token_type=ETH&direction=asc&include_fees=true&order_by=buy_quantity_with_fees&page_size=48&sell_token_address=' +
                                         itemcollection+'&sell_token_type=ERC721&status=active&sell_token_name='+itemname).json()
                        for results in r['result']:
                            if itemname == results['sell']['data']['properties']['name']:
                                productname = results['sell']['data']['properties']['name']
                                pricequantity = results['buy']['data']['quantity_with_fees']
                                pricedecimal = results['buy']['data']['decimals']
                                price = int(pricequantity) / \
                                    (10 ** int(pricedecimal))
                                USDprice = price * eth
                                break

                        if USDprice > boxprice:
                            cop += percentage

                        sheet.cell(row=counter, column=1).value = itemname
                        sheet.cell(row=counter, column=2).value = round(
                            USDprice, 2)
                        sheet.cell(row=counter, column=3).value = percentage
                        calc1 = sheet.cell(row=counter, column=2).coordinate
                        calc2 = sheet.cell(row=counter, column=3).coordinate
                        sheet.cell(row=counter, column=4).value = (
                            '='+calc1+'*'+calc2)
                        counter = counter+1

                        fair_value += percentage * USDprice

            sheet.cell(row=counter+1, column=1).value = 'Current price:'
            sheet.cell(row=counter+1, column=2).value = round(boxprice, 2)

            sheet.cell(row=counter+1, column=4).value = 'Calculated fair value:'
            sheet.cell(row=counter+1, column=5).value = '=SUM(D2:D' + \
                (str(counter-1))+')/100'
            dims = {}
            for row in sheet.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max(
                            (dims.get(cell.column_letter, 0), len(str(cell.value))))
            for col, value in dims.items():
                sheet.column_dimensions[col].width = value
            wb.save(giftbox+".xlsx")
            excel2img.export_img(giftbox+".xlsx", giftbox+".png")
            wb.close()
            embed = discord.Embed(title=boxname)
            file = discord.File(giftbox+".png", filename=giftbox+".png")
            embed.set_image(url="attachment://"+giftbox+".png")
            embed.set_thumbnail(url=image)
            embed.add_field(name="Fair value:", value='$' +
                            (str(round(fair_value/100, 2))), inline=False)
            embed.add_field(name="Actual value:", value='$' +
                            (str(round(boxprice, 2))), inline=False)
            embed.add_field(name="Profit chance:", value=(
                str(round(cop, 1))) + '%', inline=False)
            await message.channel.send(embed=embed, files=[discord.File(giftbox+'.xlsx'), file])
            # await message.channel.send(file=discord.File(giftbox+'.png'))
            # await message.channel.send('The fair value is: $' + (str(round(fair_value/100, 2))) + '\nThe actual value is: $' + (str(round(boxprice,2))) +'\nAnd the profit chance is: ' + (str(round(cop,1))) + '%', file=discord.File(giftbox+'.xlsx'))
        else:
            giftboxstring = ''
            for singlebox in giftboxarr:
                giftboxstring += singlebox + ' \n'
            await message.channel.send('**Use one of these:** \n'+giftboxstring)

    # level
    if message.content.lower().startswith(',level'):
        print(message.content)
        address = message.content.split(' ', 1)[1]
        if address == "dylan":
            address = '0xadc493644308c46a1f91f1b5aec8d2b7f14aa8aa'
        if ".eth" in address:
            r = requests.get(
                'https://api.ensideas.com/ens/resolve/'+address).json()
            try:
                address = r['address']
            except:
                await message.channel.send("You tried an invalid ENS")
        r = requests.get(
            'https://collectibles.habbo.com/api/collectibles/scoreDistribution?walletAddress='+address).json()
        xp = r['scoreDistribution']['totalScore']
        level = "error"
        max_level = 25
        level_thresholds = [0, 100, 500, 1000, 2000, 3500, 5000, 7500, 10000, 12500, 15000, 20000, 30000, 55000, 90000, 135000,
                            # Add more thresholds as needed
                            225000, 350000, 550000, 800000, 1100000, 1500000, 2000000, 2750000, 3750000, 5000000]
        next_level = 0
        level = 0
        for idx, threshold in enumerate(level_thresholds):
            if xp < threshold:
                next_level = threshold - xp
                break
            level = min(idx, max_level)
        embed = discord.Embed()
        embed.add_field(name="Level:", value=(str(level))+'/25', inline=True)
        embed.add_field(name="XP:", value=(str(xp)), inline=True)
        embed.add_field(name="Next level:", value=(
            str(next_level)) + ' XP', inline=False)
        await message.channel.send(embed=embed)

    # selling
    if message.content.lower().startswith(',selling'):
        print(message.content)
        collection = message.content.split(' ', 2)[1].lower()
        if collection in collections:
            collectionadd = collections[collection]
        address = message.content.split(' ', 2)[2]
        if address == "dylan":
            address = '0xadc493644308c46a1f91f1b5aec8d2b7f14aa8aa'
        if ".eth" in address:
            r = requests.get(
                'https://api.ensideas.com/ens/resolve/'+address).json()
            try:
                address = r['address']
            except:
                await message.channel.send("You tried an invalid ENS")
        r = requests.get(
            'https://api.x.immutable.com/v1/assets?collection='+collectionadd+'&user='+address+'&sell_orders=true&include_fees=true&order_by=updated_at&page_size=1000').json()
        ethr = requests.get(
            'https://min-api.cryptocompare.com/data/price?fsym=ETH&tsyms=USD').json()
        eth = ethr['USD']
        for shit in r['result']:
            if 'orders' in shit:
                name = shit['name']
                price = shit['orders']['sell_orders'][0]['buy_quantity']
                price = int(price) / (10 ** 18)
                r = requests.get('https://api.x.immutable.com/v3/orders?auxiliary_fee_percentages=1&auxiliary_fee_recipients=0x12cB8E42c7ec27d30df6Cb8f44aa6445D0e1a78C&buy_token_type=ETH&direction=asc&include_fees=true&order_by=buy_quantity_with_fees&page_size=200&sell_token_address=' +
                                 collectionadd+'&sell_token_type=ERC721&status=active&sell_token_name='+name).json()
                for results in r['result']:
                    if name == results['sell']['data']['properties']['name']:
                        currentprice = results['buy']['data']['quantity_with_fees']
                        currentprice = int(currentprice) / (10 ** 18)
                        price = price * eth
                        currentprice = currentprice * eth
                        pricediff = ((price-currentprice) / currentprice) * 100
                        if pricediff > 0:
                            await message.channel.send('> '+name+'\nyour listing: '+str(round(price, 2))+'\ncurrent price: '+str(round(currentprice, 2))+'\ndifference: '+str(round(pricediff, 2))+'%')
                        break

    # Secret commands?
    # p/l download
    if message.content.lower() == ',pl' and (message.author.id == 842118554878214154 or message.author.id == 128930535887208448):
        print(message.content)
        nftitems = {}
        nftshop = requests.get(
            'https://collectibles.habbo.com/api/shop/items/').json()
        ethprice = requests.get(
            'https://min-api.cryptocompare.com/data/price?fsym=ETH&tsyms=USD').json()
        eth = ethprice['USD']

        # gets the current price per credit
        emeraldamounts = ['Emerald Stone (10 Emeralds)', 'Emerald Bowl (50 Emeralds)', '100 Carat Emerald', 'Sack of 500 Emeralds',
                          '1000 Carat Emerald', 'Emerald Dino Egg (2,000 Emeralds)', 'Chest of 10k Emeralds']
        total = 0
        alllisted = 0
        for emeraldamount in emeraldamounts:
            r = requests.get('https://api.x.immutable.com/v3/orders?auxiliary_fee_percentages=1&auxiliary_fee_recipients=0x12cB8E42c7ec27d30df6Cb8f44aa6445D0e1a78C&buy_token_type=ETH&direction=asc&include_fees=true&order_by=buy_quantity_with_fees&page_size=48&sell_token_address=0xfbf1c1c09a94fe45ea8cc981c478816963ec958c&sell_token_type=ERC721&status=active&sell_token_name='+emeraldamount).json()
            for results in r['result']:
                if emeraldamount == results['sell']['data']['properties']['name']:
                    pricequantity = results['buy']['data']['quantity_with_fees']
                    pricedecimal = results['buy']['data']['decimals']
                    price = int(pricequantity) / (10 ** int(pricedecimal))
                    USDprice = price * eth
                    if emeraldamount == 'Chest of 10k Emeralds':
                        tokens = 10000
                    else:
                        tokens = int(
                            ''.join(filter(str.isdigit, emeraldamount)))
                    ppc = round(USDprice/tokens, 3)
                    total = total + ppc
                    alllisted = alllisted + 1
                    break
        total = (total/alllisted)

        # pushes the products from the NFT shop into the nftitems dictionary
        for nfti in nftshop['items']:
            if nfti['collection'] in nftitems.keys():
                nftitems[nfti['collection']][nfti['name']] = nfti['mintCost']
            else:
                nftitems[nfti['collection']] = {}
                nftitems[nfti['collection']][nfti['name']] = nfti['mintCost']

        # loop for every nft item category
        for category in nftitems:
            print(category)
            # basic file building
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.cell(row=1, column=1).value = "Name"
            sheet.cell(row=1, column=2).value = "Dollar Price"
            sheet.cell(row=1, column=3).value = "Credit Price"
            sheet.cell(row=1, column=4).value = "P/L"
            sheet['A1'].fill = PatternFill("solid", start_color="c4c2c2")
            sheet['B1'].fill = PatternFill("solid", start_color="c4c2c2")
            sheet['C1'].fill = PatternFill("solid", start_color="c4c2c2")
            sheet['D1'].fill = PatternFill("solid", start_color="c4c2c2")

            sorting = {}

            # getting item prices and pushing them to sorting
            for item in nftitems[category]:
                r = requests.get('https://api.x.immutable.com/v3/orders?auxiliary_fee_percentages=1&auxiliary_fee_recipients=0x12cB8E42c7ec27d30df6Cb8f44aa6445D0e1a78C&buy_token_type=ETH&direction=asc&include_fees=true&order_by=buy_quantity_with_fees&page_size=48&sell_token_address=' +
                                 collections[category]+'&sell_token_type=ERC721&status=active&sell_token_name='+item).json()
                for results in r['result']:
                    if item in results['sell']['data']['properties']['name']:
                        pricequantity = results['buy']['data']['quantity_with_fees']
                        pricedecimal = results['buy']['data']['decimals']
                        price = int(pricequantity) / (10 ** int(pricedecimal))
                        usdprice = price * eth
                        pl = (usdprice - (total * nftitems[category][item]))/(
                            total * nftitems[category][item]) * 100
                        sorting[item] = {"name": item, "DPrice": round(
                            usdprice, 2), "CPrice": nftitems[category][item], "PL": (round(pl, 2))}
                        break

            # sorting the sorting dictionary on P/L. (sorteddict only returns the sortitem name)
            counter = 2
            sorteddict = sorted(sorting, reverse=True,
                                key=lambda x: (sorting[x]['PL']))
            for sortitem in sorteddict:
                sheet.cell(
                    row=counter, column=1).value = sorting[sortitem]['name']
                sheet.cell(
                    row=counter, column=2).value = sorting[sortitem]['DPrice']
                sheet.cell(
                    row=counter, column=3).value = sorting[sortitem]['CPrice']
                sheet.cell(
                    row=counter, column=4).value = sorting[sortitem]['PL']
                counter = counter+1

            # resizing the rows. Yes I stole this from stackoverflow.
            dims = {}
            for row in sheet.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max(
                            (dims.get(cell.column_letter, 0), len(str(cell.value))))
            for col, value in dims.items():
                sheet.column_dimensions[col].width = value

            # saving the excel file
            wb.save(category+".xlsx")
            excel2img.export_img(category+".xlsx", category+".png")
            wb.close()
            # await message.channel.send(files=[discord.File(category+'.xlsx'),discord.File(category+'.png')])
            embed = discord.Embed(title=category)
            file = discord.File(category+".png", filename=category+".png")
            embed.set_image(url="attachment://"+category+".png")
            await message.channel.send(embed=embed, files=[discord.File(category+'.xlsx'), file])

    if message.content.lower() == ',nftc':
        # await message.channel.send("https://media.discordapp.net/attachments/1009458091328282654/1009495941709058178/unknown.png?width=822&height=702")
        await message.channel.send("https://media.discordapp.net/attachments/1188288872342360105/1294724115219087432/emeralds.png")
    if message.content.lower() == ',factory':
        await message.channel.send("https://media.discordapp.net/attachments/1111460291679617054/1125917836691374212/info_craft.png?width=451&height=701")
        await message.channel.send("https://media.discordapp.net/attachments/1111460291679617054/1125917967763390627/1QKBCzTqthodBp-0jfnC7TA.png?width=451&height=701")
    if message.content.lower() == ',crate':
        await message.channel.send("https://media.discordapp.net/attachments/1157924474344181820/1163117725044969612/image.png?ex=653e68cb&is=652bf3cb&hm=4a864146e1436bba46197ccc63225e4dedfacf1de48290466c7c20c3cc226ca3&=&width=784&height=701")
    if message.content.lower() == ',nickreset':
        await message.guild.me.edit(nick="Dark Habbo")
    if message.content.lower() == ',ping':
        await message.channel.send(message.author.id)
    if message.content.lower() == ',dylan':
        rand = random.random()
        if rand < 0.001:  # 0.1% chance
            link_message = await message.channel.send("You found a legendary Dylan! 0.1% https://media.discordapp.net/attachments/999637575818039376/1247304055315890276/boobhat.jpg?ex=665f89ec&is=665e386c&hm=f372a9854fb964155f06be6ca223496226768f26a948f287e96d5b66f0d9e28d&=&format=webp&width=702&height=936")
        elif rand < 0.101:  # 10% chance
            link_message = await message.channel.send("You found a rare Dylan! 10% \nhttps://media.discordapp.net/attachments/999637575818039376/1247301859761193034/IMG-20230605-WA0004.jpg?ex=665f87e1&is=665e3661&hm=9c58fe36dbe94597fc808176bd0452829da0576a0dd1a7f80061c4c2dc2669bc&=&format=webp&width=526&height=936")
        else:  # Remaining 89.9% chance
            link_message = await message.channel.send("https://media.discordapp.net/attachments/955933694189789234/1139988782645522432/IMG_20230812_123619.jpg?width=526&height=701")

        # anti bully protection
        command_message_id = message.id
        sleep_duration = random.randint(5, 15)
        time.sleep(sleep_duration)

        try:
            command_message = await message.channel.fetch_message(command_message_id)
        except discord.NotFound:
            command_message = None

        if command_message is None or command_message.content != message.content:
            await link_message.delete()

    if message.content.lower() == ',event':
        await message.channel.send("https://images.habbo.com/web_images/habbo-web-articles/Summer_Collectables_Promo.png")
    if message.content.lower() == ',disclevel' and message.guild.id == 1141935847743172639:
        try:
            r = requests.get(
                'https://mee6.xyz/api/plugins/levels/leaderboard/877906589846220860')
            players_json = r.json()['players']
            # player_level = [i for i in players_json if i['id'] == str(message.author.id)][0]['level']
            for index, users in enumerate(players_json):
                if users['id'] == str(message.author.id):
                    player_level = users['level']
                    player_rank = index + 1
            await message.channel.send("You are level " + (str(player_level)) + " in the habbo nft server. #" + (str(player_rank)))
        except Exception as e:
            print(e)
            await message.channel.send("Are you sure that you have a level?")
    if message.content.lower() == ',json':
        nftitems = {}
        nftshop = requests.get(
            'https://collectibles.habbo.com/api/shop/items/').json()
        for nfti in nftshop['items']:
            if nfti['collection'] not in nftitems.keys():
                nftitems[nfti['collection']] = {}

            nftitems[nfti['collection']][nfti['name']] = {}
            nftitems[nfti['collection']][nfti['name']
                                         ]['productCode'] = nfti['productCode']
            nftitems[nfti['collection']][nfti['name']
                                         ]['minted'] = nfti['minted']
            nftitems[nfti['collection']][nfti['name']
                                         ]['mintcost'] = nfti['mintCost']
            nftitems[nfti['collection']][nfti['name']
                                         ]['releaseDatetime'] = nfti['startsAtTimestamp']
            nftitems[nfti['collection']][nfti['name']
                                         ]['rarity'] = nfti['rarity']
            nftitems[nfti['collection']][nfti['name']]['ltd'] = nfti['limited']
            if nfti['collection'] == 'furniture':
                nftitems[nfti['collection']][nfti['name']
                                             ]['image'] = 'https://nft-tokens.habbo.com/items/images/'+nfti['image_url']
            else:
                nftitems[nfti['collection']][nfti['name']]['image'] = 'https://nft-tokens.habbo.com/' + \
                    nfti['collection']+'/images/'+nfti['image_url']

        with open("items.json", "w") as fp:
            json.dump(nftitems, fp)

        await message.channel.send(file=discord.File('items.json'))


@client.event
async def on_ready():
    print('Logged in as')
    print(client.user.name)
    print('------')
    ppcloop.start()

while True:
    client.run(TOKEN)
